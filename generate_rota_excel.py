from datetime import datetime, timedelta
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill
import random
random.seed(42) # Maintain reproducibility across runs

# -------------------------------
# 1️⃣ Define people, two-week work schedules
# -------------------------------
# 0=Mon, 1=Tue, ..., 4=Fri
# Each person has 'week1' and 'week2' keys
# TO HANNAH - the spelling of the person must be consistent throughout. And the number of people must also remain consistent

people = [
    "Alice","Bob","Charlie","Diana","Ethan","Fiona",
    "George","Hannah","Ian","Jane","Karl","Liam",
    "Mona","Nina","Oliver","Paula"
]

# TO HANNAH - if anyone doesn't work Fridays then they don't get Fridays on call. If there is anyone in this category then they will get 0 here in the Friday count. To discuss. 

work_schedule = {
    "Alice":    {"week1": [1,2,3,4],     "week2": [0,2,3,4]},    # Alternate Mon/Tuesday off
    "Bob":      {"week1": [0,1,2,3],   "week2": [0,2,3,4]},  # Alternate Friday, Tuesday off
    "Charlie":  {"week1": [1,2,3,4],   "week2": [0,1,2,4]},  # Alternate Monday, Thursday off
    "Diana":    {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Ethan":    {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Fiona":    {"week1": [0,1,2,3],     "week2": [0,1,2,3]},
    "George":   {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Hannah":   {"week1": [0,1,2],       "week2": [0,1,2]},
    "Ian":      {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Jane":     {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Karl":     {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Liam":     {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Mona":     {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Nina":     {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Oliver":   {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
    "Paula":    {"week1": [0,1,2,3,4],   "week2": [0,1,2,3,4]},
}

#TO HANNAH - ensure everyone is allocated to one site here

# Groups
southmead_group = ["Alice","Bob","Charlie","Diana","Ethan","Fiona","Karl","Liam"]
uhbw_group = ["George","Hannah","Ian","Jane","Mona","Nina","Oliver","Paula"]

#TO HANNAH - only those who can't swap need to be added here

# Cannot swap weekend site
cannot_swap_weekend_site = ["Alice","Charlie","George","Mona"]

#TO HANNAH - only those in ACF need to be added here

# ACF - needed for reducing their on call frequency
acf_doctors = ["Fiona", "Hannah", "Liam", "Nina"]

# Example unavailable dates

unavailable_dates = {
    "Alice": {
        ("2026-03-10", "2026-03-14"): "Annual Leave",  #TO HANNAH - this is how to do date ranges
        "2026-05-15": "Conference" #TO HANNAH - this is how to do single dates.
    },
    "Bob": {
        "2026-04-22": "Training",
        ("2026-09-05", "2026-09-06"): "Weekend Away"
    },
    "Charlie": {
        ("2026-07-01", "2026-07-03"): "Annual Leave"
    },
    "Diana": {
        ("2026-06-03", "2026-06-07"): "Annual Leave"
    },
    "Ethan": {
        "2026-02-18": "Study Leave"
    },
    "Fiona": {
        "2026-07-12": "Personal Leave",
        ("2026-11-20", "2026-11-22"): "Conference"
    },
    "George": {
        ("2026-05-01", "2026-05-03"): "Annual Leave"
    },
    "Hannah": {
        ("2026-02-25", "2026-02-28"): "Sick Leave"
    },
    "Ian": {
        "2026-06-10": "Training"
    },
    "Jane": {
        "2026-04-08": "Conference",
        ("2026-12-23", "2026-12-27"): "Christmas Leave"
    },
    "Karl": {
        ("2026-03-20", "2026-03-22"): "Annual Leave"
    },
    "Liam": {
        ("2026-08-15", "2026-08-19"): "Annual Leave"
    },
    "Mona": {
        "2026-10-03": "Wedding"
    },
    "Nina": {
        ("2026-05-10", "2026-05-12"): "Sick Leave"
    },
    "Oliver": {
        ("2026-01-02", "2026-01-04"): "New Year Leave"
    },
    "Paula": {
        ("2026-09-14", "2026-09-18"): "Annual Leave"
    },
}

#Define a function to handle the dat ranges:

def expand_unavailable(unavailable):
    expanded = {}
    for person, absences in unavailable.items():
        expanded[person] = {}
        for key, reason in absences.items():
            if isinstance(key, tuple):  # a (start, end) range
                start = datetime.strptime(key[0], "%Y-%m-%d")
                end = datetime.strptime(key[1], "%Y-%m-%d")
                current = start
                while current <= end:
                    expanded[person][current.strftime("%Y-%m-%d")] = reason
                    current += timedelta(days=1)
            else:  # a single date
                expanded[person][key] = reason
    return expanded

#Use function to expand date ranges to dates for use in rostering

unavailable = expand_unavailable(unavailable_dates)


# -------------------------------
# 2️⃣ Define rota dates
# -------------------------------
# Bank holidays 

#TO HANNAH - update this to reflect the days BH cover is needed

bank_holidays = {
    "2026-01-01": "New Year's Day",
    "2026-04-03": "Good Friday",
    "2026-04-06": "Easter Monday",
    "2026-05-04": "Early May Bank Holiday",
    "2026-05-25": "Spring Bank Holiday",
    "2026-08-31": "Summer Bank Holiday",
    "2026-12-25": "Christmas Day",
    "2026-12-28": "Boxing Day (substitute)"
}

#TO HANNAH - this is start/end date

start_date = datetime(2026, 2, 4)
end_date = datetime(2026, 8, 5)

#TO HANNAH - below here nothing should need to be edited.

all_dates = []
current_day = start_date
while current_day <= end_date:
    all_dates.append(current_day)
    current_day += timedelta(days=1)

# -------------------------------
# 3️⃣ FTE calculation
# -------------------------------
# Average FTE over two weeks and multiply 0.75 if ACF

fte = {
    p: (
        ((len(work_schedule[p]["week1"]) + len(work_schedule[p]["week2"])) / (2 * 5))
        * (0.75 if p in acf_doctors else 1)
    )
    for p in people
}

total_fte = sum(fte.values())

# -------------------------------
# 4️⃣ Weekend allocation
# -------------------------------
weekend_blocks = [(d, d + timedelta(days=1)) for d in all_dates if d.weekday()==5 and (d+timedelta(days=1))<=end_date]
total_weekends = len(weekend_blocks)*2
target_weekends = {p: total_weekends*(fte[p]/total_fte) for p in people}

weekend_assigned_southmead = defaultdict(int)
weekend_assigned_uhbw = defaultdict(int)
weekend_rota = {}

for sat, sun in weekend_blocks:
    available_sm = [p for p in southmead_group
                    if all(d.strftime("%Y-%m-%d") not in unavailable.get(p,{}) for d in [sat,sun])]
    available_uh = [p for p in uhbw_group
                    if all(d.strftime("%Y-%m-%d") not in unavailable.get(p,{}) for d in [sat,sun])]
    if not available_sm:
        available_sm = [p for p in uhbw_group if p not in cannot_swap_weekend_site
                        and all(d.strftime("%Y-%m-%d") not in unavailable.get(p,{}) for d in [sat,sun])]
    if not available_uh:
        available_uh = [p for p in southmead_group if p not in cannot_swap_weekend_site
                        and all(d.strftime("%Y-%m-%d") not in unavailable.get(p,{}) for d in [sat,sun])]
    chosen_sm = min(
        available_sm, key=lambda x: (
            weekend_assigned_southmead[x]/fte[x],
            random.random()) #If tied make it random so that those earliest in an alphabetical list are not disadvantaged
        )
    chosen_uh = min(
        available_uh, key=lambda x: (
            weekend_assigned_uhbw[x]/fte[x],
            random.random()
            )
    )

    weekend_rota[sat.strftime("%Y-%m-%d")] = {"Southmead": chosen_sm, "UHBW": chosen_uh}
    weekend_rota[sun.strftime("%Y-%m-%d")] = {"Southmead": chosen_sm, "UHBW": chosen_uh}
    weekend_assigned_southmead[chosen_sm] += 1
    weekend_assigned_uhbw[chosen_uh] += 1

# Weekend protection - no on call in the week before/after. I think this protects around BH also but can check
weekend_protection = defaultdict(set)
for sat, sun in weekend_blocks:
    weekend_dates = [sat, sun]
    for i in range(1,5):
        d_before = sat - timedelta(days=i)
        if d_before.weekday() < 5:
            weekend_dates.append(d_before)
    for i in range(1,6):
        d_after = sun + timedelta(days=i)
        if d_after.weekday() < 5:
            weekend_dates.append(d_after)
    for p in [weekend_rota[sat.strftime("%Y-%m-%d")]["Southmead"],
              weekend_rota[sat.strftime("%Y-%m-%d")]["UHBW"]]:
        weekend_protection[p].update(d.strftime("%Y-%m-%d") for d in weekend_dates)


# -------------------------------
# 5a️⃣ Bank holiday allocation (separate from weekdays)
# -------------------------------
bank_holiday_rota = {}
bank_holiday_assigned_sm = defaultdict(int)
bank_holiday_assigned_uhbw = defaultdict(int)

total_bh = len(bank_holidays)*2
target_bh = {p: total_bh*(fte[p]/total_fte) for p in people}

for bh_date_str, bh_name in bank_holidays.items():
    bh_date = datetime.strptime(bh_date_str, "%Y-%m-%d")
    
    # Southmead
    available_sm = [p for p in southmead_group if bh_date_str not in unavailable.get(p, {})
                    and bh_date_str not in weekend_protection.get(p,set())]
    # Use tie-breaker: fewest BH shifts / FTE, then fewest weekend shifts / FTE, then random to avoid systematic bias
    chosen_sm = min(
        available_sm, 
        key=lambda x: (bank_holiday_assigned_sm[x]/fte[x],
                       (weekend_assigned_southmead[x] + weekend_assigned_uhbw.get(x,0))/fte[x],
                       random.random()
        )
    )

    # UHBW
    available_uh = [p for p in uhbw_group if bh_date_str not in unavailable.get(p, {})]
    chosen_uh = min(
        available_uh, 
        key=lambda x: (bank_holiday_assigned_uhbw[x]/fte[x],
                       (weekend_assigned_uhbw[x] + weekend_assigned_southmead.get(x,0))/fte[x],
                       random.random()
        )
    )
    
    bank_holiday_rota[bh_date_str] = {"Southmead": chosen_sm, "UHBW": chosen_uh}
    bank_holiday_assigned_sm[chosen_sm] += 1
    bank_holiday_assigned_uhbw[chosen_uh] += 1


#-------------
# 5.7 Friday allocation
#---------------

total_fridays = sum(1 for d in all_dates if d.weekday() == 4)
full_time_target_friday = total_fridays / total_fte
target_fridays = {p: full_time_target_friday*fte[p] for p in people}

assigned_friday_counts = defaultdict(int)
rota = {}

for d in all_dates:
    if d.weekday() != 4:
        continue #Allocating here only Fridays
    date_str = d.strftime("%Y-%m-%d")
    if date_str in bank_holidays:
        continue  # Skip assigning on bank holidays
    week_num = ((d - start_date).days // 7) % 2
    week_key = "week1" if week_num == 0 else "week2"

    available_people = [
        p for p in people
        if (
            d.weekday() in work_schedule[p][week_key]
            and assigned_friday_counts[p] < target_fridays[p]
            and date_str not in unavailable.get(p,{})
            and date_str not in weekend_protection.get(p,set())
        )
    ]

    if not available_people:
        working_people = [
            p for p in people
            if d.weekday() in work_schedule[p][week_key]
            and date_str not in unavailable.get(p,{})
            and date_str not in weekend_protection.get(p,set())
        ]
        available_people = sorted(working_people, key=lambda x: assigned_friday_counts[x]/fte[x])

    chosen = min(available_people, 
                 key=lambda x: (assigned_friday_counts[x]/fte[x],
                                      (bank_holiday_assigned_uhbw[x] + bank_holiday_assigned_sm.get(x,0))/fte[x],
                                       (weekend_assigned_uhbw[x] + weekend_assigned_southmead.get(x,0))/fte[x],
                                       random.random()
                                )
    )
   
    rota[date_str] = chosen
    assigned_friday_counts[chosen] += 1


# -------------------------------
# 5️⃣ Weekday allocation (two-week rolling)
# -------------------------------
total_weekdays = sum(1 for d in all_dates if d.weekday() < 4)
full_time_target = total_weekdays / total_fte
target_shifts = {p: full_time_target*fte[p] for p in people}

assigned_weekday_counts = defaultdict(int)

for d in all_dates:
    if d.weekday() >= 4:
        continue #ie Mon-Thurs only
    date_str = d.strftime("%Y-%m-%d")
    if date_str in bank_holidays:
        continue  # Skip assigning on bank holidays
    week_num = ((d - start_date).days // 7) % 2
    week_key = "week1" if week_num == 0 else "week2"

    available_people = [
        p for p in people
        if (
            d.weekday() in work_schedule[p][week_key]
            and assigned_weekday_counts[p] < target_shifts[p]
            and date_str not in unavailable.get(p,{})
            and date_str not in weekend_protection.get(p,set())
        )
    ]

    if not available_people:
        working_people = [
            p for p in people
            if d.weekday() in work_schedule[p][week_key]
            and date_str not in unavailable.get(p,{})
            and date_str not in weekend_protection.get(p,set())
        ]
        available_people = sorted(working_people, key=lambda x: assigned_weekday_counts[x]/fte[x])

    chosen = min(available_people, 
                 key=lambda x: (
                     assigned_weekday_counts[x]/fte[x],
                     assigned_friday_counts[x]/fte[x],
                     (bank_holiday_assigned_uhbw[x] + bank_holiday_assigned_sm.get(x,0))/fte[x],
                     (weekend_assigned_uhbw[x] + weekend_assigned_southmead.get(x,0))/fte[x],
                     random.random()
                     )
    )
    rota[date_str] = chosen
    assigned_weekday_counts[chosen] += 1

# -------------------------------
# 6️⃣ Prepare DataFrame
# -------------------------------

dates_str = [d.strftime("%Y-%m-%d") for d in all_dates]
weekday_names = [d.strftime("%A") for d in all_dates]
data = {"Date": dates_str, "Day": weekday_names}
for p in people:
    data[p] = []

for d in all_dates:
    date_str = d.strftime("%Y-%m-%d")
    
    # 1️⃣ Bank Holiday
    if date_str in bank_holiday_rota:
        day_assignment = bank_holiday_rota[date_str]
        for person in people:
            if person == day_assignment["Southmead"]:
                data[person].append("BH SM")
            elif person == day_assignment["UHBW"]:
                data[person].append("BH UHB")
            else:
                data[person].append("")
        continue  # Skip weekend/weekday allocation for BH
    
    # 2️⃣ Weekend
    elif date_str in weekend_rota:
        day_assignment = weekend_rota[date_str]
        for person in people:
            if person == day_assignment.get("Southmead"):
                data[person].append("WC SM")
            elif person == day_assignment.get("UHBW"):
                data[person].append("WC UHB")
            else:
                data[person].append("")
        continue
    
    # 3️⃣ Weekday
    assigned_person = rota.get(date_str)
    for person in people:
        if date_str in unavailable.get(person,{}):
            data[person].append(unavailable[person][date_str])
        elif person == assigned_person:
            data[person].append("On Call")
        else:
            data[person].append("")


df = pd.DataFrame(data)
excel_filename = "on_call_rota_16people_colored.xlsx"
df.to_excel(excel_filename, index=False)

# -------------------------------
# 7️⃣ Color Excel
# -------------------------------
color_map = {
    "On Call": "90EE90",
    "WC SM": "ADD8E6",
    "WC UHB": "FFFF99",
    "Annual Leave": "FF6347",
    "Sick Leave": "FF6347",
    "Conference": "FF6347",
    "Training": "FF6347",
    "Personal Leave": "FF6347"
}

wb = openpyxl.load_workbook(excel_filename)
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
    for cell in row:
        fill_color = color_map.get(cell.value)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

wb.save(excel_filename)
print(f"Colored rota saved to '{excel_filename}'")

# -------------------------------
# 8️⃣ Print summaries
# -------------------------------

summary_rows = []

for p in people:
    fte_pct = round(fte[p] * 100)
    acf_label = "ACF" if p in acf_doctors else ""

    total_weekends_p = weekend_assigned_southmead.get(p, 0) + weekend_assigned_uhbw.get(p, 0)
    total_bh_p = bank_holiday_assigned_sm.get(p, 0) + bank_holiday_assigned_uhbw.get(p, 0)
    total_friday_p = assigned_friday_counts.get(p, 0)
    total_weekday_p = assigned_weekday_counts.get(p, 0)

    # ---- per-person total actuals ----
    total_actual = (
         total_weekends_p*2
       + total_bh_p
       + total_friday_p
       + total_weekday_p
    )

    # ---- per-person total targets ----
    total_target = (
         target_weekends[p]*2
       + target_bh[p]
       + target_fridays[p]
       + target_shifts[p]
    )

    summary_rows.append({
        "Person": p,
        "FTE / ACF": f"{fte_pct}% {acf_label}",

        "Weekends": f"{total_weekends_p} (Target: {target_weekends[p]:.1f})",
        "Bank Holidays": f"{total_bh_p} (Target: {target_bh[p]:.1f})",
        "Fridays": f"{total_friday_p} (Target: {target_fridays[p]:.1f})",
        "Weekdays": f"{total_weekday_p} (Target: {target_shifts[p]:.1f})",

        # NEW ROW TOTALS
        "Total OC": total_actual,
        "Total Target": f"{total_target:.1f}"
    })

summary_df = pd.DataFrame(summary_rows)
print(summary_df)
